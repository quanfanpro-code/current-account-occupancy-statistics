from __future__ import annotations

import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import pandas as pd

from .models import ProcessConfig

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    import pyarrow  # noqa: F401
    HAS_PYARROW = True
except ImportError:
    HAS_PYARROW = False


def supports_pyarrow_dtype_backend() -> bool:
    if not HAS_PYARROW:
        return False
    major, minor, *_ = [int(part) for part in pd.__version__.split(".")[:2]]
    return (major, minor) >= (2, 0)


class ExcelRepository:
    def get_sheet_names(self, file_path: str) -> list[str]:
        path = Path(file_path)
        if not path.exists():
            return []
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            try:
                with zipfile.ZipFile(path, "r") as zf:
                    if "xl/workbook.xml" in zf.namelist():
                        with zf.open("xl/workbook.xml") as stream:
                            parser = ET.XMLParser(resolve_entities=False)
                            tree = ET.parse(stream, parser)
                        names = [elem.get("name") for elem in tree.getroot().iter() if elem.tag.endswith("sheet") and elem.get("name")]
                        if names:
                            return names
            except Exception:
                pass
            if openpyxl is not None:
                workbook = openpyxl.load_workbook(path, read_only=True, keep_links=False)
                try:
                    return list(workbook.sheetnames)
                finally:
                    workbook.close()
        return list(pd.ExcelFile(path).sheet_names)

    def get_columns(self, file_path: str, sheet_name: str) -> list[str]:
        path = Path(file_path)
        if path.suffix.lower() in {".xlsx", ".xlsm"} and openpyxl is not None:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                if sheet_name not in workbook.sheetnames:
                    return []
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows(values_only=True):
                    if any(cell is not None and str(cell).strip() for cell in row):
                        return [str(value) if value is not None else f"Unnamed: {index}" for index, value in enumerate(row)]
                return []
            finally:
                workbook.close()
        return list(pd.read_excel(path, sheet_name=sheet_name, nrows=0).columns)

    def read_dataframes(self, config: ProcessConfig) -> tuple[pd.DataFrame, pd.DataFrame]:
        # openpyxl 仅支持 .xlsx/.xlsm，旧版 .xls 直接给出明确提示
        for path in (config.balance_file, config.trans_file):
            suffix = Path(path).suffix.lower()
            if suffix not in {".xlsx", ".xlsm"}:
                raise ValueError(f"不支持的文件格式: {suffix or path}，请另存为 .xlsx 后重试")
        read_options = {"engine": "openpyxl"}
        if supports_pyarrow_dtype_backend():
            read_options["dtype_backend"] = "pyarrow"
        try:
            df_balance = pd.read_excel(config.balance_file, sheet_name=config.balance_sheet, **read_options)
            df_trans = pd.read_excel(
                config.trans_file,
                sheet_name=config.trans_sheet,
                usecols=[config.col_name, config.col_debit, config.col_credit, config.col_date],
                **read_options,
            )
        except TypeError:
            fallback = {"engine": "openpyxl"}
            df_balance = pd.read_excel(config.balance_file, sheet_name=config.balance_sheet, **fallback)
            df_trans = pd.read_excel(
                config.trans_file,
                sheet_name=config.trans_sheet,
                usecols=[config.col_name, config.col_debit, config.col_credit, config.col_date],
                **fallback,
            )
        return df_balance, df_trans

    def save_results(self, output_path: str, data_map: dict[str, pd.DataFrame], progress_callback) -> None:
        # 调用 make_excel 生成深海蓝主题的摩根系标准报表（数据从 B2 开始）
        try:
            from make_excel import make_excel
        except ImportError:
            make_excel = None
        # 仅写入非空 DataFrame，按 ReportBundle 的 sheet 顺序输出
        sheets = [(name, df) for name, df in data_map.items() if df is not None and not df.empty]
        total = len(sheets) if sheets else 1
        if not sheets:
            # 没有任何数据时，至少创建一个空表
            sheets = [("空表", pd.DataFrame())]
        for index, (sheet_name, _) in enumerate(sheets, start=1):
            progress_callback(50 + int(index / total * 45), f"正在写入: {sheet_name}", "写入数据阶段")
        if make_excel:
            make_excel(sheets, output_path, theme="deep-navy")
        else:
            # 回退：make_excel 不可用时用裸 to_excel
            import pandas as _pd
            with _pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for name, sheet_df in sheets:
                    sheet_df.to_excel(writer, sheet_name=name, index=False)
