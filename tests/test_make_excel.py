import pandas as pd
from openpyxl import load_workbook

from make_excel import make_excel


def test_make_excel_all_empty_sheets_still_saves(tmp_path):
    """全部 sheet 为空时也必须保存出合法 xlsx，不能因没有可见 sheet 崩溃"""
    out = tmp_path / "empty.xlsx"
    make_excel([("空表", pd.DataFrame())], str(out))
    wb = load_workbook(out)
    assert len(wb.sheetnames) >= 1
